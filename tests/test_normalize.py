"""Tester for normaliseringsmodulen."""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.ingest.normalize import (
    normalize,
    skriv_country_summary_endring,
    skriv_country_summary_nok,
    skriv_tidsserier_maanedlig,
)
from src.ingest.parse_kiel import (
    BILATERAL_ARK,
    COUNTRY_SUMMARY_ARK,
    DISBURSEMENT_ARK,
    FORVENTEDE_BILATERAL_KOLONNER,
    FORVENTEDE_SUMMARY_KOLONNER,
    Aktivitet,
    LandSummary,
)
from src.analyze.tidsserier import aggreger_per_maaned


def _lag_fixture(tmp_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    bil = wb.create_sheet(BILATERAL_ARK)
    bil.append(list(FORVENTEDE_BILATERAL_KOLONNER))
    bil.append(["Norway", date(2024, 1, 1), "Military", "Allocation", 1_000_000_000])
    bil.append(["Norway", date(2024, 2, 1), "Financial", "Commitment", 500_000_000])

    cs = wb.create_sheet(COUNTRY_SUMMARY_ARK)
    for _ in range(7):
        cs.append([])
    cs.append(list(FORVENTEDE_SUMMARY_KOLONNER))
    cs.append(["Norway", 0, 1, 2.0, 1.8, 6.2, 10.0, 5.0, 1.7, 18.0, 24.7])

    disb = wb.create_sheet(DISBURSEMENT_ARK)
    for _ in range(9):
        disb.append([])
    disb.append(["", "", "", "January", "February"])
    disb.append(["", "Country", "EU member", "1", "2"])
    disb.append([])
    disb.append(["", "Norway", 0, 0.0, 0.3])

    path = tmp_path / "fixture.xlsx"
    wb.save(str(path))
    return path


def test_normalize_produserer_forventede_filer(tmp_path: Path) -> None:
    xlsx = _lag_fixture(tmp_path)
    out_dir = tmp_path / "processed"
    metadata = normalize(xlsx, out_dir=out_dir)

    summary_path = out_dir / "country_summary.csv"
    bilateral_path = out_dir / "bilateral_activities.csv"
    metadata_path = out_dir / "metadata.json"

    assert summary_path.exists()
    assert bilateral_path.exists()
    assert metadata_path.exists()

    with summary_path.open() as fh:
        rader = list(csv.DictReader(fh))
    assert len(rader) == 1
    assert rader[0]["land"] == "Norway"
    assert float(rader[0]["total_allocation"]) == pytest.approx(10.0)

    with bilateral_path.open() as fh:
        brader = list(csv.DictReader(fh))
    assert len(brader) == 2
    assert brader[0]["giver"] == "Norway"
    assert brader[0]["dato"] == "2024-01-01"
    assert float(brader[0]["verdi_eur_mrd"]) == pytest.approx(1.0)

    assert metadata["antall_land_summary"] == 1
    assert metadata["antall_bilateral_rader"] == 2
    assert metadata["antall_disbursement_rader"] == 1
    assert "sha256" in metadata

    with (out_dir / "financial_disbursements.csv").open() as fh:
        drader = list(csv.DictReader(fh))
    assert len(drader) == 1
    assert drader[0]["giver"] == "Norway"
    assert float(drader[0]["verdi_eur_mrd"]) == pytest.approx(0.3)


def _lag_fixture_med_total(tmp_path: Path, filnavn: str, total: float) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    bil = wb.create_sheet(BILATERAL_ARK)
    bil.append(list(FORVENTEDE_BILATERAL_KOLONNER))
    bil.append(["Norway", date(2024, 1, 1), "Military", "Allocation", 1_000_000_000])
    cs = wb.create_sheet(COUNTRY_SUMMARY_ARK)
    for _ in range(7):
        cs.append([])
    cs.append(list(FORVENTEDE_SUMMARY_KOLONNER))
    cs.append(["Norway", 0, 1, 2.0, 1.8, total - 3.8, total, 5.0, 1.7, 18.0, 24.7])
    disb = wb.create_sheet(DISBURSEMENT_ARK)
    for _ in range(9):
        disb.append([])
    disb.append(["", "", "", "January"])
    disb.append(["", "Country", "EU member", "1"])
    disb.append([])
    disb.append(["", "Norway", 0, 0.0])
    path = tmp_path / filnavn
    wb.save(str(path))
    return path


def test_endring_csv_skrives_med_to_releaser(tmp_path: Path) -> None:
    raw = tmp_path / "raw"
    raw.mkdir()
    _lag_fixture_med_total(raw, "2026-01-01_release27.xlsx", total=9.0)
    ny = _lag_fixture_med_total(raw, "2026-04-01_release28.xlsx", total=10.5)
    ut = tmp_path / "endring.csv"
    antall, forrige = skriv_country_summary_endring(ny, raw, ut)
    assert antall == 1
    assert forrige == "2026-01-01_release27.xlsx"
    with ut.open() as fh:
        rader = list(csv.DictReader(fh))
    assert rader[0]["land"] == "Norway"
    assert float(rader[0]["delta_total_allocation"]) == pytest.approx(1.5)


def test_endring_csv_hoppes_over_med_bare_en_release(tmp_path: Path) -> None:
    raw = tmp_path / "raw"
    raw.mkdir()
    eneste = _lag_fixture_med_total(raw, "2026-04-01_release28.xlsx", total=10.0)
    ut = tmp_path / "endring.csv"
    antall, forrige = skriv_country_summary_endring(eneste, raw, ut)
    assert antall == 0
    assert forrige is None
    assert not ut.exists()


def _finn_ekte_fil() -> Path | None:
    repo_root = Path(__file__).resolve().parents[1]
    filer = sorted((repo_root / "data" / "raw" / "kiel").glob("*.xlsx"))
    return filer[-1] if filer else None


def test_skriv_country_summary_nok_summerer_per_land(tmp_path: Path) -> None:
    aktiviteter = [
        Aktivitet("Norway", date(2024, 1, 15), "Military", "Allocation", 1_000_000),
        Aktivitet("Norway", date(2024, 1, 15), "Financial", "Commitment", 500_000),
        Aktivitet("Sweden", None, "Humanitarian", "Allocation", 200_000),
    ]
    summary = [
        LandSummary("Norway", False, True, 0, 0, 0, 0, 0, 0, 0, 0),
        LandSummary("Sweden", True, True, 0, 0, 0, 0, 0, 0, 0, 0),
    ]
    kurser = {date(2024, 1, 15): 11.20, date(2024, 4, 1): 11.50}
    out = tmp_path / "country_summary_nok.csv"
    antall = skriv_country_summary_nok(aktiviteter, summary, kurser, out)
    assert antall == 2
    with out.open() as fh:
        rader = {r["land"]: r for r in csv.DictReader(fh)}
    assert float(rader["Norway"]["military_allocation_nok"]) == pytest.approx(
        1_000_000 * 11.20
    )
    assert float(rader["Norway"]["total_allocation_nok"]) == pytest.approx(
        1_000_000 * 11.20
    )
    assert float(rader["Norway"]["financial_commitment_nok"]) == pytest.approx(
        500_000 * 11.20
    )
    # Sweden: aktivitet uten dato - skal bruke siste kurs (11.50)
    assert float(rader["Sweden"]["humanitarian_allocation_nok"]) == pytest.approx(
        200_000 * 11.50
    )


def test_skriv_tidsserier_maanedlig_skriver_csv(tmp_path: Path) -> None:
    aktiviteter = [
        Aktivitet("Norway", date(2024, 1, 15), "Military", "Allocation", 1_000_000),
        Aktivitet("Norway", date(2024, 1, 20), "Military", "Allocation", 500_000),
        Aktivitet("Norway", None, "Military", "Allocation", 9_000_000),
    ]
    kurser = {date(2024, 1, 15): 11.20}
    rader = aggreger_per_maaned(aktiviteter, kurser)
    out = tmp_path / "tidsserier_maanedlig.csv"
    skriv_tidsserier_maanedlig(rader, out)
    with out.open() as fh:
        csv_rader = list(csv.DictReader(fh))
    assert len(csv_rader) == 1
    assert csv_rader[0]["land"] == "Norway"
    assert int(csv_rader[0]["aar"]) == 2024
    assert int(csv_rader[0]["maaned"]) == 1
    assert float(csv_rader[0]["sum_eur"]) == pytest.approx(1_500_000)


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_normalize_norges_total_allocation(tmp_path: Path) -> None:
    fil = _finn_ekte_fil()
    assert fil is not None
    normalize(fil, out_dir=tmp_path)
    with (tmp_path / "country_summary.csv").open() as fh:
        for rad in csv.DictReader(fh):
            if rad["land"] == "Norway":
                assert float(rad["total_allocation"]) == pytest.approx(10.0, rel=0.05)
                return
    pytest.fail("Norge ikke funnet i country_summary.csv")
