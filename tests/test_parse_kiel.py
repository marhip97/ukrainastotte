"""Enhetstester for Kiel-parser.

Fixture-baserte tester replikerer strukturen i Release 28:
bilateral-arket (header på rad 1) og Country Summary (header på rad 8,
preamble over).

En integrasjonstest nederst kjører kun hvis ekte Kiel-fil finnes under
`data/raw/kiel/`, og krysssjekker Norges total-allocation mot kjent verdi.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.ingest.parse_kiel import (
    BILATERAL_ARK,
    COUNTRY_SUMMARY_ARK,
    DISBURSEMENT_ARK,
    EU_AID_SHARES_ARK,
    FORVENTEDE_BILATERAL_KOLONNER,
    FORVENTEDE_EU_AID_KOLONNER,
    FORVENTEDE_SUMMARY_KOLONNER,
    KolonneKontraktFeil,
    LandSummary,
    parse_bilateral,
    parse_country_summary,
    parse_eu_aid_shares,
    parse_financial_disbursements,
    parse_gdp,
    validate_summary,
)


def _lag_fixture(tmp_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    bil = wb.create_sheet(BILATERAL_ARK)
    bil.append(list(FORVENTEDE_BILATERAL_KOLONNER))
    # Kolonner: donor, announcement_date, aid_type_general, measure,
    #           tot_sub_activity_value_EUR, tot_sub_activity_value_EUR_redistr,
    #           tot_activity_value_EUR, month, month_exists_dummy
    # Jan 2022 = month 1, Jan 2024 = month 25.
    bil.append(["Norway", date(2024, 1, 15), "Military", "Allocation",
                100_000_000, 95_000_000, 110_000_000, 25, 1])
    bil.append(["Norway", date(2024, 2, 10), "Military", "Commitment",
                500_000_000, 0, 500_000_000, 26, 1])
    bil.append(["Norway", date(2024, 3, 1), "Humanitarian", "Allocation",
                20_000_000, 20_000_000, 20_000_000, 27, 1])
    # Trailing whitespace i kategori - skal normaliseres bort.
    bil.append(["Germany", date(2024, 4, 1), "Humanitarian ", "Allocation",
                50_000_000, 48_000_000, 50_000_000, 28, 1])

    cs = wb.create_sheet(COUNTRY_SUMMARY_ARK)
    for _ in range(7):
        cs.append([])
    # Etter de 11 obligatoriske summary-kolonnene har Kiel også GDP (2021) i
    # USD (kol 12) og EUR (kol 13). Headeren har "GDP (2021)" på begge.
    cs.append(list(FORVENTEDE_SUMMARY_KOLONNER) + ["GDP (2021)", "GDP (2021)"])
    cs.append(["Norway", 0, 1, 2.03, 1.78, 6.19, 10.00, 5.03, 1.66, 18.03,
               24.72, 482.0, 407.7])
    cs.append(["Germany", 1, 1, 1.0, 2.0, 15.0, 18.0, 2.0, 3.0, 20.0, 25.0,
               4260.0, 3601.7])

    path = tmp_path / "fixture.xlsx"
    wb.save(str(path))
    return path


def test_bilateral_parser_leser_rader(tmp_path: Path) -> None:
    rader = parse_bilateral(_lag_fixture(tmp_path))
    # Rader med trailing whitespace i kategori skal normaliseres og komme med.
    assert [(r.giver, r.kategori, r.maal) for r in rader] == [
        ("Norway", "Military", "Allocation"),
        ("Norway", "Military", "Commitment"),
        ("Norway", "Humanitarian", "Allocation"),
        ("Germany", "Humanitarian", "Allocation"),
    ]
    assert rader[0].dato == date(2024, 1, 15)
    assert rader[0].verdi_eur == pytest.approx(100_000_000)


def test_bilateral_parser_leser_m7_kolonner(tmp_path: Path) -> None:
    """M7.1: redistr-, activity- og maaneds-feltene leses inn."""
    rader = parse_bilateral(_lag_fixture(tmp_path))
    norge_alloc = rader[0]
    assert norge_alloc.verdi_eur_redistr == pytest.approx(95_000_000)
    assert norge_alloc.verdi_eur_activity == pytest.approx(110_000_000)
    assert norge_alloc.maaned_nr == 25
    assert norge_alloc.maaned_finnes is True
    # Commitment-rad har 0 i redistr, 500m i activity (matcher Kiels mønster).
    norge_comm = rader[1]
    assert norge_comm.verdi_eur_redistr == pytest.approx(0.0)
    assert norge_comm.verdi_eur_activity == pytest.approx(500_000_000)
    assert norge_comm.maaned_finnes is True


def test_country_summary_parser_leser_rader(tmp_path: Path) -> None:
    rader = parse_country_summary(_lag_fixture(tmp_path))
    assert [r.land for r in rader] == ["Norway", "Germany"]
    norge = rader[0]
    assert norge.er_eu_medlem is False
    assert norge.er_geografisk_europa is True
    assert norge.total_allocation == pytest.approx(10.00)
    assert norge.total_commitment == pytest.approx(24.72)


def test_parser_kaster_ved_kolonneavvik(tmp_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    bil = wb.create_sheet(BILATERAL_ARK)
    bil.append(["donor", "announcement_date", "aid_type_general"])  # mangler to
    bil.append(["Norway", date(2024, 1, 1), "Military"])
    path = tmp_path / "trunkert.xlsx"
    wb.save(str(path))
    with pytest.raises(KolonneKontraktFeil):
        parse_bilateral(path)


def _lag_disbursement_fixture(tmp_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    ark = wb.create_sheet(DISBURSEMENT_ARK)
    # Rad 1-9 kan være blanke/tittel
    for _ in range(9):
        ark.append([])
    # Rad 10: månedsetiketter
    ark.append(
        ["", "", "", "January", "February", "March", "January (2023)", "February (2023)"]
    )
    # Rad 11: header
    ark.append(["", "Country", "EU member", "1", "2", "3", "4", "5"])
    # Rad 12: tom
    ark.append([])
    # Data
    ark.append(["", "Norway", 0, 0.0, 0.5, 0.0, 0.0, 0.3])
    ark.append(["", "Germany", 1, 1.0, 0.0, 0.0, 2.0, 0.0])
    ark.append(["", "Total per month", 0, 1.0, 0.5, 0.0, 2.0, 0.3])
    path = tmp_path / "disbursement.xlsx"
    wb.save(str(path))
    return path


def _lag_eu_aid_fixture(tmp_path: Path) -> Path:
    """Fixture for EU Aid Shares - rad 1-8 preamble, rad 9 header, data fra rad 10."""
    wb = Workbook()
    wb.remove(wb.active)
    ark = wb.create_sheet(EU_AID_SHARES_ARK)
    for _ in range(8):
        ark.append([])
    # Rad 9 (header) - leading None i kol A, så Country er kol B
    ark.append(
        [None] + list(FORVENTEDE_EU_AID_KOLONNER)
    )
    # Rader 10+ - data per medlemsland
    ark.append([None, "Germany", 28064, 0.2279, 42.74, 18.52])
    ark.append([None, "France", 23689, 0.1924, 36.08, 15.63])
    ark.append([None, "Sweden", 4100, 0.0333, 6.24, 2.71])
    ark.append([None, "Total", 123000, 1.0, 187.6, 81.3])  # skal filtreres
    path = tmp_path / "eu_aid.xlsx"
    wb.save(str(path))
    return path


def test_parse_eu_aid_shares_leser_andeler(tmp_path: Path) -> None:
    rader = parse_eu_aid_shares(_lag_eu_aid_fixture(tmp_path))
    # "Total" skal filtreres bort av _er_aggregat_rad.
    assert [r.land for r in rader] == ["Germany", "France", "Sweden"]
    tyskland = rader[0]
    assert tyskland.andel == pytest.approx(0.2279)
    assert tyskland.share_committed_eur_mrd == pytest.approx(42.74)
    assert tyskland.share_allocated_eur_mrd == pytest.approx(18.52)


def test_parse_eu_aid_shares_kaster_ved_kolonneavvik(tmp_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ark = wb.create_sheet(EU_AID_SHARES_ARK)
    for _ in range(8):
        ark.append([])
    ark.append([None, "Country", "Bidrag"])  # mangler tre kolonner
    ark.append([None, "Germany", 100])
    path = tmp_path / "trunkert_eu.xlsx"
    wb.save(str(path))
    with pytest.raises(KolonneKontraktFeil):
        parse_eu_aid_shares(path)


def test_parse_gdp_leser_eur_kolonnen(tmp_path: Path) -> None:
    from src.ingest.parse_kiel import parse_gdp as parse_gdp_fn  # local for clarity
    rader = parse_gdp_fn(_lag_fixture(tmp_path))
    assert [r.land for r in rader] == ["Norway", "Germany"]
    norge = rader[0]
    # Fixture har GDP USD=482.0 og GDP EUR=407.7 - vi skal lese EUR-kolonnen.
    assert norge.bnp_2021_eur_mrd == pytest.approx(407.7)
    tyskland = rader[1]
    assert tyskland.bnp_2021_eur_mrd == pytest.approx(3601.7)


def test_disbursement_parser_leser_og_filtrerer(tmp_path: Path) -> None:
    rader = parse_financial_disbursements(_lag_disbursement_fixture(tmp_path))
    # "Total per month" skal filtreres. 0-verdier hoppes over.
    # Norway: (2022-02, 0.5), (2023-02, 0.3)
    # Germany: (2022-01, 1.0), (2023-01, 2.0)
    assert [(r.giver, r.aar, r.maaned, r.verdi_eur_mrd) for r in rader] == [
        ("Norway", 2022, 2, 0.5),
        ("Norway", 2023, 2, 0.3),
        ("Germany", 2022, 1, 1.0),
        ("Germany", 2023, 1, 2.0),
    ]
    assert rader[2].er_eu_medlem is True
    assert rader[0].er_eu_medlem is False


def _finn_ekte_fil() -> Path | None:
    repo_root = Path(__file__).resolve().parents[1]
    filer = sorted((repo_root / "data" / "raw" / "kiel").glob("*.xlsx"))
    return filer[-1] if filer else None


def _land(**endringer) -> LandSummary:
    base = dict(
        land="Test",
        er_eu_medlem=False,
        er_geografisk_europa=True,
        financial_allocation=1.0,
        humanitarian_allocation=2.0,
        military_allocation=3.0,
        total_allocation=6.0,
        financial_commitment=2.0,
        humanitarian_commitment=3.0,
        military_commitment=5.0,
        total_commitment=10.0,
    )
    base.update(endringer)
    return LandSummary(**base)


def test_validate_gyldig_rad_gir_ingen_advarsler() -> None:
    assert validate_summary([_land()]) == []


def test_validate_fanger_komponent_avvik() -> None:
    # total_allocation er 6.0 men komponenter summerer til 7.0
    rad = _land(military_allocation=4.0)
    advarsler = validate_summary([rad])
    assert any("total_allocation" in a for a in advarsler)


def test_validate_fanger_commitment_under_allocation() -> None:
    rad = _land(total_commitment=5.0, military_commitment=0.0)  # 5 < 6
    advarsler = validate_summary([rad])
    assert any("total_commitment" in a and "< total_allocation" in a for a in advarsler)


def test_validate_fanger_negativ_og_urimelig_hoeg() -> None:
    neg = _land(
        financial_allocation=-1.0,
        total_allocation=4.0,
    )
    hoey = _land(
        military_allocation=500.0,
        total_allocation=503.0,
        military_commitment=500.0,
        total_commitment=505.0,
    )
    advarsler = validate_summary([neg, hoey])
    assert any("negativ" in a for a in advarsler)
    assert any("urimelig høy" in a for a in advarsler)


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_validate_ekte_fil_kjente_avvik() -> None:
    """Validerer at ekte fil kun har kjente, lavsignaturs advarsler.

    Release 28 har små commitment<allocation-avvik for noen få land
    (sannsynligvis fordi allocation-tall er oppdatert nyere enn
    commitment). Vi aksepterer inntil 5 slike advarsler og ingen andre
    typer. Vekker testen oppmerksomhet hvis nye avvik dukker opp ved
    neste release.
    """
    fil = _finn_ekte_fil()
    assert fil is not None
    advarsler = validate_summary(parse_country_summary(fil))
    assert len(advarsler) <= 5
    for a in advarsler:
        assert "total_commitment" in a and "< total_allocation" in a, (
            f"Uventet advarselstype: {a}"
        )


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_disbursement_ekte_fil_norges_sum_under_commitment() -> None:
    """Sum av Norges månedlige utbetalinger <= Norges total_commitment.

    Dette er den logiske kryssjekken: faktiske utbetalinger kan ikke
    overstige det som er committed. Gir rask sanity-check på tidsseriene.
    """
    fil = _finn_ekte_fil()
    assert fil is not None
    utbetalinger = parse_financial_disbursements(fil)
    summary = {r.land: r for r in parse_country_summary(fil)}
    norges_utbetalt = sum(
        r.verdi_eur_mrd for r in utbetalinger if r.giver == "Norway"
    )
    norges_commitment = summary["Norway"].total_commitment
    assert norges_utbetalt <= norges_commitment
    # Også positivt signal: vi forventer minst noen utbetalinger.
    assert len(utbetalinger) > 50


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_ekte_fil_eu_aid_shares_har_kjente_medlemmer() -> None:
    """M7.1: EU Aid Shares-arket har 27 EU-medlemmer med andeler som summerer
    til ~1.0. Tyskland og Frankrike skal være på topp."""
    fil = _finn_ekte_fil()
    assert fil is not None
    andeler = parse_eu_aid_shares(fil)
    assert len(andeler) >= 26  # 27 EU-medlemmer, tillater små variasjoner
    sum_andeler = sum(a.andel for a in andeler)
    assert sum_andeler == pytest.approx(1.0, rel=0.01)
    # Tyskland skal være største bidragsyter.
    andel_map = {a.land: a for a in andeler}
    assert "Germany" in andel_map
    assert andel_map["Germany"].andel == pytest.approx(0.228, rel=0.01)
    assert andel_map["Germany"].share_allocated_eur_mrd > 10  # mrd EUR


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_ekte_fil_gdp_norges_2021_bnp() -> None:
    """M7.1: Norges BNP 2021 i EUR fra Country Summary skal være rundt 408 mrd
    (verifisert mot fasit: BNP-andel 2,4542 % × 10,005 mrd = 407,7 mrd)."""
    fil = _finn_ekte_fil()
    assert fil is not None
    bnp_rader = parse_gdp(fil)
    bnp_map = {r.land: r.bnp_2021_eur_mrd for r in bnp_rader}
    assert "Norway" in bnp_map
    assert bnp_map["Norway"] == pytest.approx(407.7, rel=0.01)
    # Tyskland skal være rundt 3601 mrd EUR.
    assert bnp_map["Germany"] == pytest.approx(3601.7, rel=0.01)


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_ekte_fil_bilateral_har_redistr_og_month() -> None:
    """M7.1: Aktivitet-objekter fra ekte fil har redistr-verdi og month-felter."""
    fil = _finn_ekte_fil()
    assert fil is not None
    aktiviteter = parse_bilateral(fil)
    # Norges total redistr-allokering skal matche fasit 10,005 mrd EUR.
    norge_alloc = sum(
        a.verdi_eur_redistr for a in aktiviteter
        if a.giver == "Norway" and a.maal == "Allocation"
    )
    assert norge_alloc / 1e9 == pytest.approx(10.005, rel=0.01)
    # Minst noen aktiviteter skal ha maaned_finnes=True.
    med_maaned = [a for a in aktiviteter if a.maaned_finnes]
    assert len(med_maaned) > 100


@pytest.mark.skipif(
    _finn_ekte_fil() is None, reason="ingen ekte Kiel-fil i data/raw/kiel/"
)
def test_krysssjekk_norges_total_allocation() -> None:
    """Kryssjekk: Norges total-allocation i Country Summary matcher
    summen av sub-aktiviteter med measure=Allocation i bilateral-arket.

    Kiel-verdier i summary er i € mrd; bilateral er i EUR. Vi tillater
    1 % toleranse fordi summary kan inkludere justeringer.
    """
    fil = _finn_ekte_fil()
    assert fil is not None
    summary = {r.land: r for r in parse_country_summary(fil)}
    bilateral = parse_bilateral(fil)
    norge_sum_eur = sum(
        r.verdi_eur for r in bilateral if r.giver == "Norway" and r.maal == "Allocation"
    )
    norge_sum_mrd = norge_sum_eur / 1_000_000_000
    expected = summary["Norway"].total_allocation
    assert norge_sum_mrd == pytest.approx(expected, rel=0.01)
