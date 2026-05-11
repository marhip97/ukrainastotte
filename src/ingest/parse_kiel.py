"""Parser for Kiel Ukraine Support Tracker XLSX (Release 28-struktur).

Leser to ark:

- **Bilateral Assistance, MAIN DATA** (langformat, én rad per sub-aktivitet).
  Har kolonner `donor`, `aid_type_general` (Financial/Humanitarian/Military),
  `measure` (Allocation eller Commitment), `tot_sub_activity_value_EUR`,
  m.fl. Se `docs/kartlegging-kiel.md` for full oversikt.
- **Country Summary (€)** (pre-aggregert per land, verdier i € mrd).
  Header ligger på rad 8. Kolonner: Country, EU member, Geographic Europe,
  Financial/Humanitarian/Military allocations og commitments,
  Total bilateral allocations/commitments.

Begge kan leses uavhengig. `parse_country_summary` gir raskt aggregerte
tall for hovedvisninger i dashboardet; `parse_bilateral` brukes når
analysemodulen trenger detaljer (f.eks. splitt på type_specific eller
månedlig tidsserie).

Designprinsipp: eksplisitte kolonnekontrakter. Hvis forventede kolonner
mangler, kastes `KolonneKontraktFeil` slik at workflow kan opprette
datakilde-Issue (R1).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

BILATERAL_ARK = "Bilateral Assistance, MAIN DATA"
COUNTRY_SUMMARY_ARK = "Country Summary (€)"
COUNTRY_SUMMARY_HEADER_RAD = 8
DISBURSEMENT_ARK = "Financial disb per Month (€)"
DISBURSEMENT_MAANED_RAD = 10
DISBURSEMENT_HEADER_RAD = 11
DISBURSEMENT_FORSTE_AAR = 2022  # Første kolonne uten årsuffix antas å være 2022
EU_AID_SHARES_ARK = "EU Aid Shares"
EU_AID_SHARES_HEADER_RAD = 9

# Posisjonsbaserte kolonner i Country Summary - GDP-kolonnene har duplikatnavn
# (kol 12 = USD, kol 13 = EUR). Vi bruker EUR-kolonnen.
COUNTRY_SUMMARY_GDP_EUR_KOL = 13
# Excel-metoden bruker Kiels pre-aggregerte EU-fordeling for medlemsland.
# Disse kolonnene ligger ved siden av i Country Summary (€).
COUNTRY_SUMMARY_EU_ALLOC_KOL = 31  # "Share in total EU allocations" (€ mrd)
COUNTRY_SUMMARY_EU_COMM_KOL = 18  # "Share in total EU commitments" (€ mrd)

MAANED_NAVN = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
}

FORVENTEDE_BILATERAL_KOLONNER: tuple[str, ...] = (
    "donor",
    "announcement_date",
    "aid_type_general",
    "measure",
    "tot_sub_activity_value_EUR",
    "tot_sub_activity_value_EUR_redistr",
    "tot_activity_value_EUR",
    "month",
    "month_exists_dummy",
)

FORVENTEDE_EU_AID_KOLONNER: tuple[str, ...] = (
    "Country",
    "Contribution to EU budget (€ million)",
    "Country Share on total EU budget",
    "Share of EU Committed aid (€ billion)",
    "Share of EU Allocated aid (€ billion)",
)

FORVENTEDE_SUMMARY_KOLONNER: tuple[str, ...] = (
    "Country",
    "EU member",
    "Geographic Europe",
    "Financial allocations",
    "Humanitarian allocations",
    "Military allocations",
    "Total bilateral allocations",
    "Financial commitments",
    "Humanitarian commitments",
    "Military commitments",
    "Total bilateral commitments",
)

KATEGORIER = ("Financial", "Humanitarian", "Military")
MAAL = ("Allocation", "Commitment")


class KolonneKontraktFeil(ValueError):
    """Kolonnestruktur i Kiel-XLSX matcher ikke forventet kontrakt."""


@dataclass(frozen=True)
class Aktivitet:
    """Én sub-aktivitet fra bilateral-arket.

    Verdifelter (alle i EUR, ikke mrd):
    - `verdi_eur`: `tot_sub_activity_value_EUR` (legacy rå-verdi; fjernes i M7.3
      når dashboardet er fullt migrert til excel-metoden).
    - `verdi_eur_redistr`: `tot_sub_activity_value_EUR_redistr` (sub-aktivitetnivå
      med EU-redistribusjon - hovedverdi for allokering i excel-metoden).
    - `verdi_eur_activity`: `tot_activity_value_EUR` (aktivitetnivå; brukes til
      å summere totale forpliktelser siden commitment-rader ofte mangler
      sub-activity-verdi).

    Månedsfelter:
    - `maaned_nr`: heltall fra 1 (Jan 2022) og oppover. None hvis ukjent.
    - `maaned_finnes`: True hvis `month_exists_dummy = 1` i rådata.
      Brukes som filter på enkeltårs-aggregeringer.
    """

    giver: str
    dato: date | None
    kategori: str
    maal: str
    verdi_eur: float
    verdi_eur_redistr: float = 0.0
    verdi_eur_activity: float = 0.0
    maaned_nr: int | None = None
    maaned_finnes: bool = False


@dataclass(frozen=True)
class EuAndel:
    """En EU-medlemslands andel av EU-institusjonenes støtte.

    Hentet fra Kiels `EU Aid Shares`-ark. Brukes i M7.2 av modulen
    `src/analyze/eu_fordeling.py` for å lage INKL EU-visninger.

    - `andel`: `Country Share on total EU budget` (desimal 0-1).
    - `share_committed_eur_mrd`: Kiels pre-aggregerte forpliktelse via EU
      (i € mrd). Excel-metoden bruker dette direkte istedenfor å reprodusere
      via andel × EU-total, jf. sak S23.
    - `share_allocated_eur_mrd`: tilsvarende for allokering.
    """

    land: str
    andel: float
    share_committed_eur_mrd: float
    share_allocated_eur_mrd: float


@dataclass(frozen=True)
class LandBnp:
    """BNP per land fra Kiels `Country Summary (€)`-ark.

    `bnp_2021_eur_mrd` = `GDP (2021)`-kolonnen i EUR (kol 13 i Country Summary,
    siden GDP-kolonnene har duplikatnavn der kol 12 er USD og kol 13 er EUR).
    Brukes som BNP-grunnlag i excel-metoden istedenfor WDI BNP, jf. F4 i M7-plan.
    """

    land: str
    bnp_2021_eur_mrd: float


@dataclass(frozen=True)
class FinanciellUtbetaling:
    """En månedlig finansiell utbetaling fra ett giverland. Verdi i € mrd."""

    giver: str
    er_eu_medlem: bool
    aar: int
    maaned: int
    verdi_eur_mrd: float


@dataclass(frozen=True)
class LandSummary:
    """Aggregat per land fra Country Summary (€). Verdier i € mrd."""

    land: str
    er_eu_medlem: bool
    er_geografisk_europa: bool
    financial_allocation: float
    humanitarian_allocation: float
    military_allocation: float
    total_allocation: float
    financial_commitment: float
    humanitarian_commitment: float
    military_commitment: float
    total_commitment: float


def _til_float(verdi) -> float:
    if verdi is None or verdi == "":
        return 0.0
    try:
        return float(verdi)
    except (TypeError, ValueError):
        return 0.0


def _til_bool(verdi) -> bool:
    return _til_float(verdi) > 0.5


def _til_dato(verdi) -> date | None:
    if isinstance(verdi, datetime):
        return verdi.date()
    if isinstance(verdi, date):
        return verdi
    return None


def _hent_header(ark, radnummer: int = 1) -> tuple[str, ...]:
    for i, row in enumerate(ark.iter_rows(values_only=True), 1):
        if i == radnummer:
            return tuple(
                str(c).strip() if c is not None else "" for c in row
            )
    return ()


def _sjekk_kontrakt(header: tuple[str, ...], forventet: tuple[str, ...]) -> None:
    mangler = [k for k in forventet if k not in header]
    if mangler:
        raise KolonneKontraktFeil(
            f"Mangler forventede kolonner: {mangler}"
        )


def parse_bilateral(xlsx_path: Path) -> list[Aktivitet]:
    """Parse bilateral-arket og returner liste over sub-aktiviteter.

    Leser både legacy `tot_sub_activity_value_EUR` og excel-metode-kolonnene
    `tot_sub_activity_value_EUR_redistr`, `tot_activity_value_EUR`, `month`
    og `month_exists_dummy`. Legacy-feltet fjernes i M7.3.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[BILATERAL_ARK]
    header = _hent_header(ark, radnummer=1)
    _sjekk_kontrakt(header, FORVENTEDE_BILATERAL_KOLONNER)
    idx = {navn: header.index(navn) for navn in FORVENTEDE_BILATERAL_KOLONNER}
    resultat: list[Aktivitet] = []
    for i, row in enumerate(ark.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        giver = row[idx["donor"]]
        if giver is None or str(giver).strip() == "":
            continue
        kategori = str(row[idx["aid_type_general"]] or "").strip()
        maal = str(row[idx["measure"]] or "").strip()
        if kategori not in KATEGORIER or maal not in MAAL:
            continue
        maaned_raa = row[idx["month"]]
        maaned_nr: int | None
        if maaned_raa is None or maaned_raa == "":
            maaned_nr = None
        else:
            try:
                maaned_nr = int(float(maaned_raa))
            except (TypeError, ValueError):
                maaned_nr = None
        maaned_finnes = _til_bool(row[idx["month_exists_dummy"]])
        resultat.append(
            Aktivitet(
                giver=str(giver).strip(),
                dato=_til_dato(row[idx["announcement_date"]]),
                kategori=kategori,
                maal=maal,
                verdi_eur=_til_float(row[idx["tot_sub_activity_value_EUR"]]),
                verdi_eur_redistr=_til_float(
                    row[idx["tot_sub_activity_value_EUR_redistr"]]
                ),
                verdi_eur_activity=_til_float(row[idx["tot_activity_value_EUR"]]),
                maaned_nr=maaned_nr,
                maaned_finnes=maaned_finnes,
            )
        )
    return resultat


def parse_eu_aid_shares(xlsx_path: Path) -> list[EuAndel]:
    """Parse `EU Aid Shares`-arket og returner andel per medlemsland.

    Header står på rad 9 (EU_AID_SHARES_HEADER_RAD). Rader fra og med 10
    inneholder ett medlemsland per rad fram til "Total" eller blank rad.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[EU_AID_SHARES_ARK]
    rader = list(ark.iter_rows(values_only=True))
    if len(rader) < EU_AID_SHARES_HEADER_RAD:
        raise KolonneKontraktFeil(
            f"{EU_AID_SHARES_ARK}: for få rader ({len(rader)})"
        )
    header_rad = rader[EU_AID_SHARES_HEADER_RAD - 1]
    header = tuple(
        str(c).strip() if c is not None else "" for c in header_rad
    )
    _sjekk_kontrakt(header, FORVENTEDE_EU_AID_KOLONNER)
    idx = {k: header.index(k) for k in FORVENTEDE_EU_AID_KOLONNER}
    resultat: list[EuAndel] = []
    for rad in rader[EU_AID_SHARES_HEADER_RAD:]:
        if len(rad) <= idx["Country"]:
            continue
        land = rad[idx["Country"]]
        if land is None or str(land).strip() == "":
            continue
        land_str = str(land).strip()
        if _er_aggregat_rad(land_str):
            continue
        resultat.append(
            EuAndel(
                land=land_str,
                andel=_til_float(rad[idx["Country Share on total EU budget"]]),
                share_committed_eur_mrd=_til_float(
                    rad[idx["Share of EU Committed aid (€ billion)"]]
                ),
                share_allocated_eur_mrd=_til_float(
                    rad[idx["Share of EU Allocated aid (€ billion)"]]
                ),
            )
        )
    return resultat


def parse_gdp(xlsx_path: Path) -> list[LandBnp]:
    """Parse BNP 2021 (i € mrd) per land fra `Country Summary (€)`.

    `GDP (2021)`-kolonnene har duplikatnavn i headeren (USD og EUR). Vi
    leser posisjonsbasert som "to kolonner etter Total bilateral commitments"
    siden Kiel-arket har strukturen: ...Total bilateral commitments,
    GDP (USD), GDP (EUR), ...
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[COUNTRY_SUMMARY_ARK]
    header = _hent_header(ark, radnummer=COUNTRY_SUMMARY_HEADER_RAD)
    _sjekk_kontrakt(header, FORVENTEDE_SUMMARY_KOLONNER)
    country_idx = header.index("Country")
    total_comm_idx = header.index("Total bilateral commitments")
    gdp_eur_idx = total_comm_idx + 2  # USD-kol + 1 = EUR-kol
    resultat: list[LandBnp] = []
    for i, row in enumerate(ark.iter_rows(values_only=True), 1):
        if i <= COUNTRY_SUMMARY_HEADER_RAD:
            continue
        if len(row) <= gdp_eur_idx:
            continue
        land = row[country_idx]
        if land is None or str(land).strip() == "":
            continue
        land_str = str(land).strip()
        if _er_aggregat_rad(land_str):
            continue
        resultat.append(
            LandBnp(
                land=land_str,
                bnp_2021_eur_mrd=_til_float(row[gdp_eur_idx]),
            )
        )
    return resultat


def _er_aggregat_rad(navn: str) -> bool:
    """Sanne navn på aggregatrader i Kiel-arkene (ikke enkeltland)."""
    lavere = navn.lower().strip()
    if lavere in {"total", "eu"}:
        return True
    if "total" in lavere or "cummulative" in lavere or "cumulative" in lavere:
        return True
    return False


def _parse_maaned_kolonner(
    maaned_rad: tuple, forste_datakolonne: int = 3
) -> list[tuple[int, int, int]]:
    """Returner liste av (kolonneindeks, år, måned) for hver datakolonne.

    Måned-raden har etiketter som "January", "February", ..., "January (2023)".
    Celler uten årsuffix antas å tilhøre `DISBURSEMENT_FORSTE_AAR` inntil
    neste årsuffix dukker opp.
    """
    resultat: list[tuple[int, int, int]] = []
    aar = DISBURSEMENT_FORSTE_AAR
    for idx, celle in enumerate(maaned_rad):
        if idx < forste_datakolonne or celle is None:
            continue
        tekst = str(celle).strip()
        if not tekst:
            continue
        maaned_del = tekst
        if "(" in tekst and tekst.endswith(")"):
            navn, rest = tekst.split("(", 1)
            maaned_del = navn.strip()
            try:
                aar = int(rest.rstrip(")").strip())
            except ValueError:
                pass
        maaned = MAANED_NAVN.get(maaned_del)
        if maaned is None:
            continue
        resultat.append((idx, aar, maaned))
    return resultat


def parse_financial_disbursements(xlsx_path: Path) -> list[FinanciellUtbetaling]:
    """Parse arket "Financial disb per Month (€)" til månedlige utbetalinger.

    Tekstverdier som "No budget support" og tomme celler hoppes over.
    Rader med 0 inkluderes ikke (redusere støy; 0 betyr ingen utbetaling
    den måneden).
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[DISBURSEMENT_ARK]
    alle_rader = list(ark.iter_rows(values_only=True))
    if len(alle_rader) < DISBURSEMENT_HEADER_RAD:
        raise KolonneKontraktFeil(
            f"{DISBURSEMENT_ARK}: for få rader ({len(alle_rader)})"
        )
    maaned_rad = alle_rader[DISBURSEMENT_MAANED_RAD - 1]
    header_rad = alle_rader[DISBURSEMENT_HEADER_RAD - 1]
    if str(header_rad[1] or "").strip() != "Country":
        raise KolonneKontraktFeil(
            f"{DISBURSEMENT_ARK}: forventet 'Country' i kolonne B rad "
            f"{DISBURSEMENT_HEADER_RAD}, fant {header_rad[1]!r}"
        )
    maaned_map = _parse_maaned_kolonner(maaned_rad)
    if not maaned_map:
        raise KolonneKontraktFeil(
            f"{DISBURSEMENT_ARK}: fant ingen månedskolonner i rad "
            f"{DISBURSEMENT_MAANED_RAD}"
        )
    resultat: list[FinanciellUtbetaling] = []
    for rad in alle_rader[DISBURSEMENT_HEADER_RAD:]:
        giver = rad[1] if len(rad) > 1 else None
        if giver is None or str(giver).strip() == "":
            continue
        giver_tekst = str(giver).strip()
        if _er_aggregat_rad(giver_tekst):
            continue
        er_eu = _til_bool(rad[2] if len(rad) > 2 else 0)
        for idx, aar, maaned in maaned_map:
            if idx >= len(rad):
                continue
            verdi = _til_float(rad[idx])
            if verdi == 0:
                continue
            resultat.append(
                FinanciellUtbetaling(
                    giver=giver_tekst,
                    er_eu_medlem=er_eu,
                    aar=aar,
                    maaned=maaned,
                    verdi_eur_mrd=verdi,
                )
            )
    return resultat


def validate_summary(
    rader: list[LandSummary],
    komponent_toleranse: float = 0.01,
    rimelig_maks_mrd: float = 200.0,
) -> list[str]:
    """Returner liste med advarsler for mistenkelige rader.

    Regler:
    - `total_allocation` skal være sum av tre komponent-allokeringer
      (tillatt avvik: `komponent_toleranse`, default 1 %).
    - `total_commitment` skal være sum av tre komponent-commitments.
    - `total_commitment >= total_allocation` (allocation er delmengde
      av commitment).
    - Ingen verdi skal være negativ.
    - Ingen verdi skal overstige `rimelig_maks_mrd` (default 200 mrd EUR).

    Tom liste betyr at alle rader passerte. Brukes av `qa`-rollen og av
    fetch-workflow for å oppdage strukturelle feil tidlig.
    """
    advarsler: list[str] = []
    for r in rader:
        alloc_komponenter = (
            r.financial_allocation
            + r.humanitarian_allocation
            + r.military_allocation
        )
        if r.total_allocation and abs(alloc_komponenter - r.total_allocation) > (
            komponent_toleranse * r.total_allocation
        ):
            advarsler.append(
                f"{r.land}: total_allocation ({r.total_allocation:.3f}) avviker "
                f"fra sum komponenter ({alloc_komponenter:.3f})"
            )
        comm_komponenter = (
            r.financial_commitment
            + r.humanitarian_commitment
            + r.military_commitment
        )
        if r.total_commitment and abs(comm_komponenter - r.total_commitment) > (
            komponent_toleranse * r.total_commitment
        ):
            advarsler.append(
                f"{r.land}: total_commitment ({r.total_commitment:.3f}) avviker "
                f"fra sum komponenter ({comm_komponenter:.3f})"
            )
        if r.total_commitment + komponent_toleranse < r.total_allocation:
            advarsler.append(
                f"{r.land}: total_commitment ({r.total_commitment:.3f}) "
                f"< total_allocation ({r.total_allocation:.3f})"
            )
        for navn, verdi in (
            ("financial_allocation", r.financial_allocation),
            ("humanitarian_allocation", r.humanitarian_allocation),
            ("military_allocation", r.military_allocation),
            ("total_allocation", r.total_allocation),
            ("financial_commitment", r.financial_commitment),
            ("humanitarian_commitment", r.humanitarian_commitment),
            ("military_commitment", r.military_commitment),
            ("total_commitment", r.total_commitment),
        ):
            if verdi < 0:
                advarsler.append(f"{r.land}: negativ {navn} = {verdi}")
            if verdi > rimelig_maks_mrd:
                advarsler.append(
                    f"{r.land}: urimelig høy {navn} = {verdi} (> {rimelig_maks_mrd})"
                )
    return advarsler


def parse_country_summary(xlsx_path: Path) -> list[LandSummary]:
    """Parse Country Summary (€) og returner aggregat per land."""
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[COUNTRY_SUMMARY_ARK]
    header = _hent_header(ark, radnummer=COUNTRY_SUMMARY_HEADER_RAD)
    _sjekk_kontrakt(header, FORVENTEDE_SUMMARY_KOLONNER)
    idx = {navn: header.index(navn) for navn in FORVENTEDE_SUMMARY_KOLONNER}
    resultat: list[LandSummary] = []
    for i, row in enumerate(ark.iter_rows(values_only=True), 1):
        if i <= COUNTRY_SUMMARY_HEADER_RAD:
            continue
        land = row[idx["Country"]]
        if land is None or str(land).strip() == "":
            continue
        # "Total"-raden og eventuelle EU-aggregater filtreres - de er
        # ikke enkeltland og gir falske advarsler i validate_summary.
        if _er_aggregat_rad(str(land)):
            continue
        resultat.append(
            LandSummary(
                land=str(land).strip(),
                er_eu_medlem=_til_bool(row[idx["EU member"]]),
                er_geografisk_europa=_til_bool(row[idx["Geographic Europe"]]),
                financial_allocation=_til_float(row[idx["Financial allocations"]]),
                humanitarian_allocation=_til_float(
                    row[idx["Humanitarian allocations"]]
                ),
                military_allocation=_til_float(row[idx["Military allocations"]]),
                total_allocation=_til_float(row[idx["Total bilateral allocations"]]),
                financial_commitment=_til_float(
                    row[idx["Financial commitments"]]
                ),
                humanitarian_commitment=_til_float(
                    row[idx["Humanitarian commitments"]]
                ),
                military_commitment=_til_float(row[idx["Military commitments"]]),
                total_commitment=_til_float(row[idx["Total bilateral commitments"]]),
            )
        )
    return resultat
