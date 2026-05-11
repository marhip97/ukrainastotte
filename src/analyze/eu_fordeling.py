"""EU-fordeling for INKL EU-visning i dashboardet.

Per sak S23: bruk Kiels pre-aggregerte INKL EU-verdier direkte fra
`Country Summary (€)`-arket istedenfor å reprodusere via fordelingsnøkkel.
Kiel publiserer "Total bilateral and EU allocations" (kol 33) og "Total
bilateral and EU commitments" (kol 20) som inkluderer hvert medlemslands
andel av EU- og EIB-bidragene. Disse tallene treffer fasit eksakt
(jf. M7.0-rapport seksjon 1).

Funksjonen `fordelEuStotte` returnerer en utvidet summary med både
EKSKL EU-verdier (uendret fra `LandSummary`) og INKL EU-verdier per land.
Ikke-EU-medlemmer har identiske EKSKL og INKL-verdier.

EU-fordelingsnøkkel (`Country Share on total EU budget`) beholdes som
referanseverdi i `EuAndel`-strukturen, men er ikke beregningsgrunnlag.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.ingest.parse_kiel import (
    COUNTRY_SUMMARY_ARK,
    COUNTRY_SUMMARY_HEADER_RAD,
    LandSummary,
    _er_aggregat_rad,
    _hent_header,
    _til_float,
)

# Posisjonsbaserte kolonner i Country Summary (€) for INKL EU-totaler.
# Verifisert mot Release 28 i M7.0-utforskning.
KOL_TOTAL_INKL_EU_COMMITMENTS = 20  # "Total bilateral and EU commitments" (€ mrd)
KOL_TOTAL_INKL_EU_ALLOCATIONS = 33  # "Total bilateral and EU allocations" (€ mrd)


@dataclass(frozen=True)
class LandSummaryEu:
    """Land-aggregat med både EKSKL og INKL EU-verdier.

    EKSKL-feltene er identiske med `LandSummary`. INKL-feltene reflekterer
    Kiels pre-aggregerte "bilateral + EU/EIB"-totaler. For ikke-EU-medlemmer
    er EKSKL og INKL identiske.
    """

    land: str
    er_eu_medlem: bool
    er_geografisk_europa: bool
    # EKSKL EU - direkte bilateral
    financial_allocation: float
    humanitarian_allocation: float
    military_allocation: float
    total_allocation: float
    financial_commitment: float
    humanitarian_commitment: float
    military_commitment: float
    total_commitment: float
    # INKL EU - bilateral + landets andel av EU- og EIB-bidrag
    total_allocation_inkl_eu: float
    total_commitment_inkl_eu: float


def parse_inkl_eu_totaler(xlsx_path: Path) -> dict[str, tuple[float, float]]:
    """Returner {land: (total_alloc_inkl_eu, total_comm_inkl_eu)} fra Kiel.

    Leser Country Summary (€) kol 20 og 33 posisjonsbasert siden kolonnenavn
    er duplikate i headeren.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ark = wb[COUNTRY_SUMMARY_ARK]
    resultat: dict[str, tuple[float, float]] = {}
    for i, row in enumerate(ark.iter_rows(values_only=True), 1):
        if i <= COUNTRY_SUMMARY_HEADER_RAD:
            continue
        max_kol = max(
            KOL_TOTAL_INKL_EU_ALLOCATIONS,
            KOL_TOTAL_INKL_EU_COMMITMENTS,
        )
        if len(row) <= max_kol:
            continue
        land = row[1]
        if land is None or str(land).strip() == "":
            continue
        land_str = str(land).strip()
        if _er_aggregat_rad(land_str):
            continue
        alloc = _til_float(row[KOL_TOTAL_INKL_EU_ALLOCATIONS])
        comm = _til_float(row[KOL_TOTAL_INKL_EU_COMMITMENTS])
        resultat[land_str] = (alloc, comm)
    return resultat


def fordelEuStotte(
    summary: list[LandSummary],
    inkl_eu_totaler: dict[str, tuple[float, float]],
) -> list[LandSummaryEu]:
    """Kombiner EKSKL-summary med Kiels pre-aggregerte INKL EU-totaler.

    For land som mangler INKL EU-data (ikke i Kiels Country Summary, eller
    EU-institusjonen selv) settes INKL = EKSKL som fornuftig fallback.
    """
    resultat: list[LandSummaryEu] = []
    for s in summary:
        inkl = inkl_eu_totaler.get(s.land)
        if inkl is None:
            inkl_alloc = s.total_allocation
            inkl_comm = s.total_commitment
        else:
            inkl_alloc, inkl_comm = inkl
        resultat.append(
            LandSummaryEu(
                land=s.land,
                er_eu_medlem=s.er_eu_medlem,
                er_geografisk_europa=s.er_geografisk_europa,
                financial_allocation=s.financial_allocation,
                humanitarian_allocation=s.humanitarian_allocation,
                military_allocation=s.military_allocation,
                total_allocation=s.total_allocation,
                financial_commitment=s.financial_commitment,
                humanitarian_commitment=s.humanitarian_commitment,
                military_commitment=s.military_commitment,
                total_commitment=s.total_commitment,
                total_allocation_inkl_eu=inkl_alloc,
                total_commitment_inkl_eu=inkl_comm,
            )
        )
    return resultat
